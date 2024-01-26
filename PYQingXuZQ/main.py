# 这是一个示例 Python 脚本。

# 按 Ctrl+F5 执行或将其替换为您的代码。
# 按 双击 Shift 在所有地方搜索类、文件、工具窗口、操作和设置。
# pyinstaller -F main.py

import requests, json
import datetime
import openpyxl
import os
import base64


def file_to_base64(file_path):
    with open(file_path, 'rb') as file:
        file_data = file.read()
        base64_data = base64.b64encode(file_data)
        return base64_data.decode()

import xlsxwriter as xw

def print_hi(name):
    # 在下面的代码行中使用断点来调试脚本。
    print(f'Hi, {name}')  # 按 F9 切换断点。

    nowdate = datetime.datetime.now()
    wkd = datetime.datetime.now().weekday()
    print(os.getcwd())   # 获得当前工作目录
    xlsxpatch = os.path.join(os.getcwd(),'晋级率模板.xlsx')
    print(xlsxpatch)
    data = openpyxl.load_workbook(xlsxpatch)
    #print(data.get_named_ranges())  # 输出工作页索引范围
    #print(data.get_sheet_names())  # 输出所有工作页的名称
    # 取第一张表
    # sheetnames = data.get_sheet_names()
    # table = data.get_sheet_by_name(sheetnames[0])
    table = data.active
    print(table.title)  # 输出表名
    nrows = table.max_row  # 获得行数
    # ncolumns = table.max_column  # 获得列数

    for index in range(0,-1,-1):
        try:
            indexDate = nowdate + datetime.timedelta(days=-index)
            print(indexDate)
            # if not is_workday(indexDate.date()):
            #     continue
            nowdateCode = indexDate.strftime('%Y%m%d')

            xingQi = indexDate.weekday() + 1
            if xingQi in {6,7}:
                continue

            ZTURL = 'http://push2ex.eastmoney.com/getTopicZTPool?cb=callbackdata5701347&ut=7eea3edcaed734bea9cbfc24409ed989&dpt=wz.ztzt&Pageindex=0&pagesize=320&sort=fbt%3Aasc&date=' + nowdateCode + '&_=1647191502623'
            ZTreq = requests.get(ZTURL)
            ZTres = json.loads(ZTreq.text[20:-2])
            if not ZTres['data']:
                continue
            ZTNum = ZTres['data']['tc']  # 涨停数
            YB = 0  # 首板数
            EB = 0  # 二板数
            SB = 0  # 三板数
            SiB = 0  # 三板数
            WuB = 0  # 四板数
            LiuB = 0  # 五板数
            QiB = 0  # 六板数
            BaB = 0  # 八板数
            GB = 0  # 高板数
            YZS = 0  # 一字板数
            LBS = 0  # 连板数
            for row in ZTres['data']['pool']:
                if row['lbc'] == 1:
                    YB = YB + 1
                if row['lbc'] == 2:
                    EB = EB + 1
                if row['lbc'] == 3:
                    SB = SB + 1
                if row['lbc'] == 4:
                    SiB = SiB + 1
                if row['lbc'] == 5:
                    WuB = WuB + 1
                if row['lbc'] == 6:
                    LiuB = LiuB + 1
                if row['lbc'] == 7:
                    QiB = QiB + 1
                if row['lbc'] == 8:
                    BaB = BaB + 1
                if row['lbc'] > 8:
                    GB = GB + 1
                if str(row['fbt'])[0:3] == '925':
                    YZS = YZS + 1
                if row['lbc'] > 1:
                    LBS = LBS + 1

            LBL = str(round((LBS / ZTNum) * 100, 2)) + '%'  # 连板率

            ZRZTURL = 'http://push2ex.eastmoney.com/getYesterdayZTPool?cb=callbackdata2371067&ut=7eea3edcaed734bea9cbfc24409ed989&dpt=wz.ztzt&Pageindex=0&pagesize=170&sort=zs%3Adesc&date=' + nowdateCode + '&_=1647245228771'
            ZRZTreq = requests.get(ZRZTURL)
            ZRZTres = json.loads(ZRZTreq.text[20:-2])
            ZRYB = 0  # 昨日首板数
            ZREB = 0  # 昨日二板数
            ZRSB = 0  # 昨日三板数
            ZRSiB = 0  # 昨日四板数
            ZRWuB = 0  # 昨日五板数
            ZRLiuB = 0  # 昨日六板数
            ZRQiB = 0  # 昨日七板数
            for row in ZRZTres['data']['pool']:
                if row['ylbc'] == 1:
                    ZRYB = ZRYB + 1
                if row['ylbc'] == 2:
                    ZREB = ZREB + 1
                if row['ylbc'] == 3:
                    ZRSB = ZRSB + 1
                if row['ylbc'] == 4:
                    ZRSiB = ZRSiB + 1
                if row['ylbc'] == 5:
                    ZRWuB = ZRWuB + 1
                if row['ylbc'] == 6:
                    ZRLiuB = ZRLiuB + 1
                if row['ylbc'] == 7:
                    ZRQiB = ZRQiB + 1

            YJEL = str(round((EB / ZRYB) * 100, 2)) + '%'  # 一进二率

            EJSL = '0.0%'
            if ZREB != 0:
                EJSL = str(round((SB / ZREB) * 100, 2)) + '%'  # 二进三率

            SJSL = '0.0%'
            if ZRSB != 0:
                SJSL = str(round((SiB / ZRSB) * 100, 2)) + '%'  # 三进四率

            SJWL = '0.0%'
            if ZRSiB != 0:
                SJWL = str(round((WuB / ZRSiB) * 100, 2)) + '%'  # 四进五率

            WJLL = '0.0%'
            if ZRWuB != 0:
                WJLL = str(round((LiuB / ZRWuB) * 100, 2)) + '%'  # 五进六率

            LJQL = '0.0%'
            if ZRLiuB != 0:
                LJQL = str(round((QiB / ZRLiuB) * 100, 2)) + '%'  # 六进七率

            QJBL = '0.0%'
            if ZRQiB != 0:
                QJBL = str(round((BaB / ZRQiB) * 100, 2)) + '%'  # 七进八率

            ZBURL = 'http://push2ex.eastmoney.com/getTopicZBPool?cb=callbackdata1795803&ut=7eea3edcaed734bea9cbfc24409ed989&dpt=wz.ztzt&Pageindex=0&pagesize=170&sort=fbt%3Aasc&date=' + nowdateCode + '&_=1647194511027'
            ZBreq = requests.get(ZBURL)
            ZBres = json.loads(ZBreq.text[20:-2])
            if not ZBres['data']:
                continue
            ZBNum = ZBres['data']['tc']  # 炸板数

            DTURL = 'http://push2ex.eastmoney.com/getTopicDTPool?cb=callbackdata4846344&ut=7eea3edcaed734bea9cbfc24409ed989&dpt=wz.ztzt&Pageindex=0&pagesize=170&sort=fund%3Aasc&date=' + nowdateCode + '&_=1647194811873'
            DTreq = requests.get(DTURL)
            DTres = json.loads(DTreq.text[20:-2])
            if not DTres['data']:
                DTNum = 0
            if DTres['data']:
                DTNum = DTres['data']['tc']  # 跌停数

            ZBL = str(round((ZBNum / (ZBNum + ZTNum) * 100), 2)) + '%'  # 炸板率
            ZDTB = ''
            if DTNum != 0:
                ZDTB = str(round((DTNum / ZTNum) * 100, 2)) + '%'  # 涨跌停比

            # ZDJSURL='https://push2.eastmoney.com/api/qt/ulist.np/get?cb=zdjs&fltt=2&secids=1.000001%2C0.399001&fields=f1%2Cf2%2Cf3%2Cf4%2Cf6%2Cf12%2Cf13%2Cf104%2Cf105%2Cf106&ut=b2884a393a59ad64002292a3e90d46a5&_=1647589508681'
            # ZDJSreq = requests.get(ZDJSURL)
            # ZDJSres = json.loads(ZDJSreq.text[5:-2])
            # SZJS=ZDJSres['data']['diff'][0]['f104'] + ZDJSres['data']['diff'][1]['f104'] #上涨家数
            # XDJS = ZDJSres['data']['diff'][0]['f105'] + ZDJSres['data']['diff'][0]['f105']  # 下跌家数
            # ZDB =  str(round((XDJS/SZJS)*100,2))+'%' #涨跌比

            print(str(nowdateCode) + ',星期' + str(xingQi)
                  + ',涨停数' + str(ZTNum)
                  + '，炸板数' + str(ZBNum)
                  + '，跌停数' + str(DTNum)
                  + '，炸板率' + str(ZBL)
                  + '，涨跌停比' + str(ZDTB)
                  + '，首板数' + str(YB)
                  + '，二板数' + str(EB)
                  + '，三板数' + str(SB)
                  + '，四板数' + str(SiB)
                  + '，五板数' + str(WuB)
                  + '，六板数' + str(LiuB)
                  + '，六板数' + str(QiB)
                  + '，八板数' + str(BaB)
                  + '，妖股数' + str(GB)
                  + '，连板数' + str(LBS)
                  + '，一进二率' + str(YJEL)
                  + '，二进三率' + str(EJSL)
                  + '，三进四率' + str(SJSL)
                  + '，四进五率' + str(SJWL)
                  + '，五进六率' + str(WJLL)
                  + '，六进七率' + str(LJQL)
                  + '，七进八率' + str(QJBL)
                  + '，一字板数' + str(YZS)
                  + '，连板率' + str(LBL))

            values = [nowdateCode,
                      str(xingQi),
                      str(ZTNum),
                      str(DTNum),
                      str(ZDTB),
                      str(ZBL),
                      str(ZBNum),
                      str(LBL),
                      str(YB),
                      str(LBS),
                      str(YZS),
                      str(YJEL),
                      str(EJSL),
                      str(SJSL),
                      str(SJWL),
                      str(WJLL),
                      str(LJQL),
                      str(QJBL),
                      str(YB),
                      str(EB),
                      str(SB),
                      str(SiB),
                      str(WuB),
                      str(LiuB),
                      str(QiB),
                      str(GB),
                      str(GB)
                      ]
            cells = 1
            nrows = nrows + 1
            for value in values:
                table.cell(nrows, cells).value = value
                cells = cells + 1

        except Exception as r:
            print(r)

    data.save(xlsxpatch)
    print("完成，excel新增了一行，请打开查看！")

    url = "https://api.github.com/repos/WolfMoss/AutoRuns/contents/PYQingXuZQ/晋级率模板.xlsx"

    payload = {}
    headers = {
        'Accept': 'application/vnd.github+json',
        'Authorization': 'Bearer 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'
    }

    response = requests.request("GET", url, headers=headers, data=payload)

    oldfilejson= response.json()

    # 将文件转换为 Base64
    base64_string = file_to_base64(xlsxpatch)

    print('Base64 编码:', base64_string)

    # GitHub API 地址
    url = 'https://api.github.com/repos/{owner}/{repo}/contents/{path}'

    # 请求头中需要包含身份验证信息
    headers = {
        'Authorization': 'Bearer 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',
        'Accept': 'application/vnd.github+json'
    }

    # 文件路径和内容
    file_path = 'PYQingXuZQ/晋级率模板.xlsx'

    # 构建请求数据
    data = {
        'message': 'Update file',
        'content': base64_string,
        'sha':oldfilejson['sha']
    }

    # 替换 URL 中的占位符
    url = url.format(owner='WolfMoss', repo='AutoRuns', path=file_path)

    # 发送 PUT 请求来更新文件
    response = requests.put(url, headers=headers, json=data)

    # 检查响应状态码
    if response.status_code == 200:
        print('文件更新成功')
    else:
        print('文件更新失败:', response.text)

# 按间距中的绿色按钮以运行脚本。
if __name__ == '__main__':
    print_hi('PyCharm')

# 访问 https://www.jetbrains.com/help/pycharm/ 获取 PyCharm 帮助
